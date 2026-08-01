"""Evaluation API routes."""

import io
import shutil
import uuid
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.config import settings
from app.core import converter, xml_parser
from app.db.session import get_db
from app.dependencies import get_current_user
from app.models.user import User
from app.schemas import EvaluationResponse, EvaluationListResponse
from app.services import evaluation_service

router = APIRouter(prefix="/evaluations", tags=["Evaluations"])


@router.post("", response_model=EvaluationResponse, status_code=201)
def create_evaluation(
    question_id: str = Form(...),
    student_name: str = Form(""),
    student_id: str = Form(""),
    pkt_file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Upload a .pkt file and run evaluation against a question."""
    try:
        evaluation = evaluation_service.run_evaluation(db, question_id, pkt_file, student_name, student_id, created_by=user.id)
        return evaluation
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        raise HTTPException(500, f"Evaluation failed: {str(e)}")


@router.get("", response_model=EvaluationListResponse)
def list_evaluations(
    question_id: str | None = None,
    student_id: str | None = None,
    passed: bool | None = None,
    latest_only: bool = True,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    # Admins see all, others see only their own
    from app.models.user import UserRole
    owner = None if user.role == UserRole.admin else user.id
    items = evaluation_service.get_evaluations(
        db, question_id, student_id, passed, created_by=owner, latest_only=latest_only
    )
    return EvaluationListResponse(items=items, total=len(items))


@router.get("/export")
def export_evaluations_excel(
    from_date: str | None = Query(None, description="Start date YYYY-MM-DD"),
    to_date: str | None = Query(None, description="End date YYYY-MM-DD"),
    question_id: str | None = Query(None, description="Filter by question ID"),
    passed: bool | None = Query(None, description="Filter by pass/fail status"),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Export evaluations as Excel file with summary + detailed per-check breakdown."""
    from app.models.evaluation import Evaluation
    from app.models.question import Question
    from app.models.user import UserRole

    query = db.query(Evaluation)

    # Admins see all, others see only their own
    if user.role != UserRole.admin:
        query = query.filter(Evaluation.created_by == user.id)

    # Filters
    if question_id:
        query = query.filter(Evaluation.question_id == question_id)
    if from_date:
        try:
            start = datetime.strptime(from_date, "%Y-%m-%d")
            query = query.filter(Evaluation.evaluated_at >= start)
        except ValueError:
            raise HTTPException(400, "Invalid from_date format. Use YYYY-MM-DD")
    if to_date:
        try:
            end = datetime.strptime(to_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            query = query.filter(Evaluation.evaluated_at <= end)
        except ValueError:
            raise HTTPException(400, "Invalid to_date format. Use YYYY-MM-DD")

    raw_evals = query.order_by(Evaluation.evaluated_at.desc()).all()

    # Deduplicate: Keep only final/latest result per candidate per question
    seen = set()
    evals = []
    for ev in raw_evals:
        key = (ev.student_id or ev.student_name or ev.created_by or "anon", ev.question_id)
        if key not in seen:
            seen.add(key)
            evals.append(ev)

    # Apply Pass/Fail filter to final results
    if passed is not None:
        evals = [ev for ev in evals if ev.passed == passed]

    # Build Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    wb = Workbook()

    # ──────────────────────────────────────────────
    # SHEET 1: Summary
    # ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="7C5CFC", end_color="7C5CFC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["#", "Roll Number", "Student Name", "Student ID", "Slot Timing", "Question", "Topic", "Score", "Max Score",
               "Percentage", "Checks Passed", "Checks Failed", "Status", "Date", "Time"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    pass_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    fail_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

    for i, ev in enumerate(evals, 1):
        question = db.query(Question).filter(Question.id == ev.question_id).first()
        q_title = question.title if question else "Unknown"
        q_topic = question.topic.name if question and question.topic else "Unknown"
        pct = (ev.overall_score / ev.max_score * 100) if ev.max_score > 0 else 0
        dt = ev.evaluated_at

        # Extract check results
        raw = ev.results or {}
        checks = raw if isinstance(raw, list) else raw.get("check_results", [])
        passed_count = sum(1 for c in checks if c.get("passed"))
        failed_count = len(checks) - passed_count

        row_data = [
            i,
            getattr(ev, 'roll_number', None) or "—",
            ev.student_name or "",
            ev.student_id or "",
            getattr(ev, 'session_slot', None) or "—",
            q_title,
            q_topic,
            round(ev.overall_score, 1),
            round(ev.max_score, 1),
            f"{pct:.1f}%",
            f"{passed_count}/{len(checks)}",
            str(failed_count),
            "PASSED" if ev.passed else "FAILED",
            dt.strftime("%Y-%m-%d") if dt else "",
            dt.strftime("%H:%M:%S") if dt else "",
        ]
        row_fill = pass_fill if ev.passed else fail_fill
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.border = thin_border
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center") if col != 2 else Alignment(horizontal="left")

    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 3, 35)

    # ──────────────────────────────────────────────
    # SHEET 2: Detailed Check Results
    # ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Detailed Results")

    detail_headers = ["#", "Student Name", "Student ID", "Question", "Check #", "Check Type",
                      "Description", "Status", "Score", "Message", "Expected", "Found"]
    for col, h in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    detail_row = 2
    for ev in evals:
        question = db.query(Question).filter(Question.id == ev.question_id).first()
        q_title = question.title if question else "Unknown"

        raw = ev.results or {}
        checks = raw if isinstance(raw, list) else raw.get("check_results", [])

        for ci, check in enumerate(checks, 1):
            details = check.get("details", {}) or {}
            expected = str(details.get("expected", "")) if details.get("expected") is not None else ""
            found = str(details.get("found", "")) if details.get("found") is not None else ""

            row_data = [
                detail_row - 1,
                ev.student_name or "",
                ev.student_id or "",
                q_title,
                ci,
                check.get("check_type", ""),
                check.get("check_description", ""),
                "PASS" if check.get("passed") else "FAIL",
                f"{check.get('score', 0) * 100:.0f}%",
                check.get("message", ""),
                expected,
                found,
            ]
            row_fill = pass_fill if check.get("passed") else fail_fill
            for col, val in enumerate(row_data, 1):
                cell = ws2.cell(row=detail_row, column=col, value=val)
                cell.border = thin_border
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal="left") if col in (7, 10, 11, 12) else Alignment(horizontal="center")
            detail_row += 1

    for col in ws2.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max_length + 3, 50)

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"evaluations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/{evaluation_id}", response_model=EvaluationResponse)
def get_evaluation(evaluation_id: str, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    ev = evaluation_service.get_evaluation(db, evaluation_id)
    if not ev: raise HTTPException(404, "Evaluation not found")
    return ev


@router.post("/preview", tags=["Debug"])
def preview_pkt_file(
    pkt_file: UploadFile = File(...),
    user: User = Depends(get_current_user),
):
    """Parse a .pkt file and return its contents WITHOUT evaluating.

    Professors use this to see what devices, interfaces, VLANs, and configs
    are inside a student's file — helpful to verify device names before
    creating the evaluation plan.
    """
    # Save the uploaded file
    pkt_dir = settings.upload_dir / "pkt"
    pkt_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{uuid.uuid4().hex[:8]}_{pkt_file.filename}"
    pkt_path = pkt_dir / filename

    with open(pkt_path, "wb") as f:
        shutil.copyfileobj(pkt_file.file, f)

    # Convert to XML
    try:
        xml_path = converter.convert_pkt_to_xml(pkt_path)
    except Exception:
        if pkt_file.filename and pkt_file.filename.endswith(".xml"):
            xml_path = pkt_path
        else:
            raise HTTPException(400, "Failed to convert .pkt file. Is it a valid Packet Tracer file?")

    # Parse and build summary
    network = xml_parser.parse_xml_file(xml_path)

    devices_summary = []
    for d in network.devices:
        ifaces = []
        for iface in d.running_config.interfaces:
            iface_info = {"name": iface.name, "commands": iface.commands[:10]}
            ifaces.append(iface_info)

        vlans = [{"id": v.number, "name": v.name} for v in d.vlans if v.number < 1002]

        ports_info = []
        for p in d.ports:
            port = {"name": p.name, "ip": p.ip, "subnet": p.subnet, "mac": p.mac_address}
            if p.ip:  # Only include ports with IPs
                ports_info.append(port)

        running_cfg_lines = []
        for iface in d.running_config.interfaces:
            running_cfg_lines.append(f"interface {iface.name}")
            for cmd in iface.commands:
                running_cfg_lines.append(f" {cmd}")

        global_cmds = d.running_config.global_commands

        devices_summary.append({
            "name": d.name,
            "type": d.device_type,
            "model": d.model,
            "interfaces": ifaces,
            "vlans": vlans,
            "ports_with_ip": ports_info,
            "global_commands": global_cmds,
            "running_config_sections": running_cfg_lines,
        })

    links_summary = []
    for link in network.links:
        from_dev = network.get_device_by_ref(link.from_device_ref)
        to_dev = network.get_device_by_ref(link.to_device_ref)
        links_summary.append({
            "from": f"{from_dev.name if from_dev else '?'}:{link.from_port}",
            "to": f"{to_dev.name if to_dev else '?'}:{link.to_port}",
            "type": link.link_type,
        })

    # Resolve device roles
    from app.core.role_resolver import resolve_roles
    role_map = resolve_roles(network)

    return {
        "version": network.version,
        "device_count": len(network.devices),
        "link_count": len(network.links),
        "device_names": [d.name for d in network.devices],
        "device_roles": role_map.to_dict(),
        "devices": devices_summary,
        "links": links_summary,
    }


@router.post("/batch", response_model=list[EvaluationResponse])
def batch_evaluate(
    question_id: str = Form(...),
    pkt_files: list[UploadFile] = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Batch evaluate multiple .pkt files against the same question."""
    results = []
    for pkt_file in pkt_files:
        try:
            # Extract student name from filename
            name = pkt_file.filename.rsplit(".", 1)[0] if pkt_file.filename else ""
            ev = evaluation_service.run_evaluation(db, question_id, pkt_file, name, "", created_by=user.id)
            results.append(ev)
        except Exception:
            pass  # Skip failed files in batch mode
    return results


