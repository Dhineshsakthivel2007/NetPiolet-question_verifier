"""Auth API routes — login, register, Google OAuth, admin user management."""

import httpx
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile, status
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.config import settings
from app.db.session import get_db
from app.dependencies import get_current_user, require_admin, require_admin_or_professor
from app.models.user import User, UserRole
from app.schemas import (
    GoogleLoginRequest, TokenResponse, UserApproveRequest,
    UserCreate, UserLogin, UserResponse, UserRoleRequest,
)
from app.services import auth_service

router = APIRouter(prefix="/auth", tags=["Authentication"])


@router.post("/register", status_code=status.HTTP_201_CREATED)
def register(data: UserCreate, db: Session = Depends(get_db)):
    """Register a new account. Account is inactive until admin approves."""
    slot_val = data.session_slot.strip() if data.session_slot else None
    if data.roll_number and db.query(User).filter(User.roll_number == data.roll_number, User.session_slot == slot_val).first():
        raise HTTPException(status_code=400, detail=f"Roll Number '{data.roll_number}' is already registered for slot '{slot_val or 'No Slot'}'.")
    if db.query(User).filter(User.email == data.email, User.session_slot == slot_val).first():
        raise HTTPException(status_code=400, detail=f"Email '{data.email}' is already registered for slot '{slot_val or 'No Slot'}'.")

    role = UserRole.student
    if data.role in ("admin", "professor"):
        role = UserRole(data.role)

    user = User(
        username=data.username,
        email=data.email,
        hashed_password=auth_service.hash_password(data.password),
        role=role,
        is_active=False,
        roll_number=data.roll_number,
        session_slot=slot_val,
        level_id=data.level_id,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return {"message": "Account created. Please wait for admin approval to login.", "user_id": user.id}


@router.post("/login", response_model=TokenResponse)
def login(data: UserLogin, db: Session = Depends(get_db)):
    """Login and receive JWT token."""
    matching_users = auth_service.authenticate_users(db, data.username, data.password)
    if not matching_users:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid credentials")

    selected_user = None
    denied_reason = None

    for user in matching_users:
        if user.role != UserRole.student:
            selected_user = user
            break

        if not user.is_active:
            denied_reason = "Your account has been deactivated or is pending admin approval."
            continue

        attendance = getattr(user, 'attendance', None) or "Absent"
        if attendance != "Present":
            denied_reason = "Login denied. Attendance not marked as Present. Please contact your instructor."
            continue

        allowed, msg = _is_slot_active(user.session_slot)
        if not allowed:
            denied_reason = msg
            continue

        selected_user = user
        break

    if not selected_user:
        selected_user = matching_users[0]
        if not selected_user.is_active:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Your account has been deactivated or is pending admin approval.")
        if selected_user.role == UserRole.student:
            attendance = getattr(selected_user, 'attendance', None) or "Absent"
            if attendance != "Present":
                raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Login denied. Attendance not marked as Present. Please contact your instructor.")
            allowed, msg = _is_slot_active(selected_user.session_slot)
            if not allowed:
                raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail=msg)

    token = auth_service.create_access_token({"sub": str(selected_user.id), "username": selected_user.username, "role": selected_user.role.value})

    # Single Active Session Enforcement: Flag dual login on student test sessions & set current token
    role_str = selected_user.role.value if hasattr(selected_user.role, 'value') else str(selected_user.role)
    if role_str == "student":
        from app.models.test_session import TestSession
        student_sessions = db.query(TestSession).filter(
            TestSession.student_id == selected_user.id
        ).all()
        for ts in student_sessions:
            ts.is_completed = True
            ts.dual_login_flag = True
            ts.completion_reason = "Dual Login Detected"
            ts.expires_at = datetime.utcnow()

    selected_user.is_active = True
    selected_user.current_session_token = token
    db.commit()

    return TokenResponse(access_token=token, role=role_str, username=selected_user.username)


import re
from datetime import datetime
from fastapi.responses import Response
from app.models.level import Level


def _is_slot_active(session_slot: str | None) -> tuple[bool, str]:
    """Verify if current time falls within assigned session slot timing (e.g. '09:00-11:00')."""
    if not session_slot or not session_slot.strip():
        return True, ""
    match = re.search(r'(\d{1,2}:\d{2})\s*-\s*(\d{1,2}:\d{2})', session_slot.strip())
    if not match:
        return True, ""
    try:
        now_time = datetime.now().time()
        start_time = datetime.strptime(match.group(1), "%H:%M").time()
        end_time = datetime.strptime(match.group(2), "%H:%M").time()
        if not (start_time <= now_time <= end_time):
            return False, f"Login denied. Your assigned slot timing is '{session_slot}'. Current time ({now_time.strftime('%H:%M')}) is outside your slot."
    except Exception:
        pass
    return True, ""


@router.post("/google", response_model=TokenResponse)
def google_login(data: GoogleLoginRequest, db: Session = Depends(get_db)):
    """Authenticate or register user via Google ID Token."""
    try:
        resp = httpx.get(f"https://oauth2.googleapis.com/tokeninfo?id_token={data.id_token}", timeout=10.0)
        if resp.status_code != 200:
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid Google token")
        payload = resp.json()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail=f"Failed to verify Google token: {str(e)}")

    email = payload.get("email", "").strip().lower()
    if not email:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Google account email not found")

    uname_prefix = email.split("@")[0].split(".")[0]  # e.g. "nandhakumara"

    # Match user by email case-insensitively, or username/email prefix match
    from sqlalchemy import func, or_
    user = db.query(User).filter(
        or_(
            func.lower(User.email) == email,
            func.lower(User.email).like(f"{uname_prefix}.%"),
            func.lower(User.email).like(f"{uname_prefix}@%"),
            func.lower(User.username).like(f"{uname_prefix}%")
        )
    ).order_by(User.created_at.asc()).first()

    if user and user.email != email:
        # Update email to actual authenticated Google email
        user.email = email
        db.commit()

    if not user:
        # Create new user for Google login
        username = payload.get("name") or email.split("@")[0]
        avatar_url = payload.get("picture", "")
        google_id = payload.get("sub", "")

        user = User(
            username=username,
            email=email,
            hashed_password="",
            role=UserRole.student,
            is_active=True,
            google_id=google_id,
            avatar_url=avatar_url,
            attendance="Absent",
        )
        db.add(user)
        db.commit()
        db.refresh(user)

    if not user.is_active:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Your account has been deactivated or is pending admin approval.")

    # Check attendance & slot timing for students only
    if user.role == UserRole.student:
        attendance = getattr(user, 'attendance', None) or "Absent"
        if attendance != "Present":
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Login denied. Attendance not marked as Present. Please contact your instructor.")

        allowed, msg = _is_slot_active(user.session_slot)
        if not allowed:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail=msg)

    token = auth_service.create_access_token({"sub": str(user.id), "username": user.username, "role": user.role.value})

    # Single Active Session Enforcement: Flag dual login on student test sessions & set current token
    role_str = user.role.value if hasattr(user.role, 'value') else str(user.role)
    if role_str == "student":
        from app.models.test_session import TestSession
        student_sessions = db.query(TestSession).filter(
            TestSession.student_id == user.id
        ).all()
        for ts in student_sessions:
            ts.is_completed = True
            ts.dual_login_flag = True
            ts.completion_reason = "Dual Login Detected"
            ts.expires_at = datetime.utcnow()

    user.is_active = True
    user.current_session_token = token
    db.commit()

    return TokenResponse(access_token=token, role=role_str, username=user.username)


@router.get("/me", response_model=UserResponse)
def get_me(current_user: User = Depends(get_current_user)):
    """Get current authenticated user."""
    res = UserResponse.model_validate(current_user)
    if current_user.level:
        res.level_name = current_user.level.name
    return res


# ---- Admin: User Management ----

@router.get("/users", response_model=list[UserResponse])
def list_users(db: Session = Depends(get_db), admin: User = Depends(require_admin_or_professor)):
    """List all users (admin or professor)."""
    users = db.query(User).order_by(User.created_at.desc()).all()
    res = []
    for u in users:
        ur = UserResponse.model_validate(u)
        if u.level:
            ur.level_name = u.level.name
        res.append(ur)
    return res


@router.put("/users/{user_id}/approve")
def approve_user(user_id: str, data: UserApproveRequest, db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    """Approve or deactivate a user (admin only)."""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if user.id == admin.id:
        raise HTTPException(status_code=400, detail="Cannot modify your own status")
    user.is_active = data.is_active
    db.commit()
    return {"message": f"User {'approved' if data.is_active else 'deactivated'}", "user_id": user.id}


@router.put("/users/{user_id}/role")
def change_role(user_id: str, data: UserRoleRequest, db: Session = Depends(get_db), admin: User = Depends(require_admin)):
    """Change a user's role (admin only)."""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    try:
        user.role = UserRole(data.role)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid role. Must be: admin, professor, student")
    db.commit()
    return {"message": f"Role changed to {data.role}", "user_id": user.id}


@router.delete("/users/{user_id}")
def delete_user(user_id: str, db: Session = Depends(get_db), admin: User = Depends(require_admin_or_professor)):
    """Delete a user and all related records (admin or professor)."""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if user.id == admin.id:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")

    try:
        from app.models.test_session import TestSession
        from app.models.evaluation import Evaluation
        from app.models.report import Report
        from app.models.question import Question

        # 1. Delete reports linked to user's evaluations
        eval_ids = [e.id for e in db.query(Evaluation.id).filter(
            (Evaluation.student_id == user_id) | (Evaluation.created_by == user_id)
        ).all()]
        if eval_ids:
            db.query(Report).filter(Report.evaluation_id.in_(eval_ids)).delete(synchronize_session=False)

        # 2. Delete evaluations (clears references to projects.id)
        db.query(Evaluation).filter(
            (Evaluation.student_id == user_id) | (Evaluation.created_by == user_id)
        ).delete(synchronize_session=False)

        # 3. Delete test sessions
        db.query(TestSession).filter(TestSession.student_id == user_id).delete(synchronize_session=False)

        # 4. Delete projects
        try:
            from app.models.project import Project
            db.query(Project).filter(Project.student_id == user_id).delete(synchronize_session=False)
        except Exception:
            pass

        # 5. Nullify questions created_by
        db.query(Question).filter(Question.created_by == user_id).update({"created_by": None}, synchronize_session=False)

        # 6. Delete user
        db.delete(user)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete user: {str(e)}")

    return {"message": f"User '{user.username}' and all related records deleted"}


class BulkUsersRequest(BaseModel):
    user_ids: list[str]


@router.post("/users/bulk-deactivate")
def bulk_deactivate_users(data: BulkUsersRequest, db: Session = Depends(get_db), admin: User = Depends(require_admin_or_professor)):
    """Deactivate multiple users at once."""
    target_ids = [uid for uid in data.user_ids if uid != str(admin.id)]
    if not target_ids:
        return {"message": "No valid users selected for deactivation"}

    count = db.query(User).filter(User.id.in_(target_ids)).update({"is_active": False}, synchronize_session=False)
    db.commit()
    return {"message": f"Successfully deactivated {count} users", "deactivated_count": count}


@router.post("/users/bulk-activate")
def bulk_activate_users(data: BulkUsersRequest, db: Session = Depends(get_db), admin: User = Depends(require_admin_or_professor)):
    """Activate multiple users at once."""
    target_ids = [uid for uid in data.user_ids if uid != str(admin.id)]
    if not target_ids:
        return {"message": "No valid users selected for activation"}

    count = db.query(User).filter(User.id.in_(target_ids)).update({"is_active": True}, synchronize_session=False)
    db.commit()
    return {"message": f"Successfully activated {count} users", "activated_count": count}


@router.post("/users/bulk-delete")
def bulk_delete_users(data: BulkUsersRequest, db: Session = Depends(get_db), admin: User = Depends(require_admin_or_professor)):
    """Delete multiple users and all related records at once."""
    target_ids = [uid for uid in data.user_ids if uid != str(admin.id)]
    if not target_ids:
        return {"message": "No valid users selected for deletion"}

    deleted_count = 0
    from app.models.test_session import TestSession
    from app.models.evaluation import Evaluation
    from app.models.report import Report
    from app.models.question import Question

    for uid in target_ids:
        user = db.query(User).filter(User.id == uid).first()
        if not user:
            continue
        try:
            eval_ids = [e.id for e in db.query(Evaluation.id).filter((Evaluation.student_id == uid) | (Evaluation.created_by == uid)).all()]
            if eval_ids:
                db.query(Report).filter(Report.evaluation_id.in_(eval_ids)).delete(synchronize_session=False)
            db.query(Evaluation).filter((Evaluation.student_id == uid) | (Evaluation.created_by == uid)).delete(synchronize_session=False)
            db.query(TestSession).filter(TestSession.student_id == uid).delete(synchronize_session=False)
            try:
                from app.models.project import Project
                db.query(Project).filter(Project.student_id == uid).delete(synchronize_session=False)
            except Exception:
                pass
            db.query(Question).filter(Question.created_by == uid).update({"created_by": None}, synchronize_session=False)
            db.delete(user)
            deleted_count += 1
        except Exception:
            db.rollback()

    db.commit()
    return {"message": f"Successfully deleted {deleted_count} users", "deleted_count": deleted_count}


class UserSlotUpdateRequest(BaseModel):
    session_slot: str


@router.put("/users/{user_id}/slot")
def update_user_slot(user_id: str, data: UserSlotUpdateRequest, db: Session = Depends(get_db), admin: User = Depends(require_admin_or_professor)):
    """Update student slot timing (admin or professor)."""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.session_slot = data.session_slot
    db.commit()
    return {"message": "Slot timing updated", "session_slot": user.session_slot}


class UserAttendanceRequest(BaseModel):
    attendance: str


@router.put("/users/{user_id}/attendance")
def update_user_attendance(user_id: str, data: UserAttendanceRequest, db: Session = Depends(get_db), admin: User = Depends(require_admin_or_professor)):
    """Update student attendance status (admin or professor)."""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    normalized = data.attendance.strip().title()  # "present" -> "Present", "ABSENT" -> "Absent"
    user.attendance = normalized if normalized in ("Present", "Absent") else "Absent"
    db.commit()
    return {"message": f"Attendance set to {user.attendance}", "user_id": user.id, "attendance": user.attendance}


class BulkAttendanceRequest(BaseModel):
    session_slot: str | None = None
    user_ids: list[str] | None = None
    attendance: str = "Present"


@router.post("/users/bulk-attendance")
def bulk_update_attendance(data: BulkAttendanceRequest, db: Session = Depends(get_db), admin: User = Depends(require_admin_or_professor)):
    """Mark attendance in bulk by slot timing or user IDs (admin or professor)."""
    normalized = data.attendance.strip().title()
    target_status = normalized if normalized in ("Present", "Absent") else "Present"
    query = db.query(User).filter(User.role == UserRole.student)

    if data.session_slot:
        query = query.filter(User.session_slot == data.session_slot)
    elif data.user_ids:
        query = query.filter(User.id.in_(data.user_ids))

    users = query.all()
    count = 0
    for u in users:
        u.attendance = target_status
        count += 1

    db.commit()
    return {"message": f"Marked {count} student(s) as {target_status}", "count": count, "attendance": target_status}


# ---- Admin: Create Single User (auto-activated) ----

class AdminUserCreate(BaseModel):
    username: str
    email: str
    password: str
    role: str = "student"
    roll_number: str | None = None
    session_slot: str | None = None
    level_id: str | None = None


@router.post("/users/create", status_code=status.HTTP_201_CREATED)
def admin_create_user(data: AdminUserCreate, db: Session = Depends(get_db), admin: User = Depends(require_admin_or_professor)):
    """Admin or Professor creates a user — account is immediately active."""
    slot_val = data.session_slot.strip() if (data.session_slot and data.session_slot.strip()) else None
    roll_val = data.roll_number.strip() if (data.roll_number and data.roll_number.strip()) else None
    email_val = data.email.strip().lower()

    if roll_val and db.query(User).filter(User.roll_number == roll_val, User.session_slot == slot_val).first():
        raise HTTPException(status_code=400, detail=f"Roll Number '{roll_val}' is already registered for slot '{slot_val or 'No Slot'}'.")
    if db.query(User).filter(User.email == email_val, User.session_slot == slot_val).first():
        raise HTTPException(status_code=400, detail=f"Email '{email_val}' is already registered for slot '{slot_val or 'No Slot'}'.")

    try:
        role = UserRole(data.role)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid role. Must be: admin, professor, student")

    user = User(
        username=data.username.strip(),
        email=email_val,
        hashed_password=auth_service.hash_password(data.password),
        role=role,
        is_active=True,
        roll_number=roll_val,
        session_slot=slot_val,
        level_id=data.level_id if (data.level_id and data.level_id.strip()) else None,
        attendance="Absent",
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return {"message": f"User '{data.username}' created and activated.", "user_id": user.id}


@router.get("/sample-template")
def download_sample_template():
    """Download sample CSV template for bulk student import."""
    csv_content = (
        "roll_number,username,email,password,role,session_slot,level_name\n"
        "7376221EC101,dhinesh,dhineshs.ad24@bitsathy.ac.in,Pass123!,student,09:00-11:00,Level 1\n"
        "7376221EC102,nandha,nandhakumars.cs24@bitsathy.ac.in,Pass123!,student,14:00-16:00,Level 2\n"
    )
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=sample_student_import.csv"}
    )


@router.get("/sample-template-excel")
def download_sample_template_excel():
    """Download sample Excel (.xlsx) template for bulk student import."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Student Import Template"

    headers = ["roll_number", "username", "email", "password", "role", "session_slot", "level_name"]
    ws.append(headers)

    sample_rows = [
        ["7376221EC101", "dhinesh", "dhineshs.ad24@bitsathy.ac.in", "Pass123!", "student", "09:00-11:00", "Level 1"],
        ["7376221EC102", "nandha", "nandhakumars.cs24@bitsathy.ac.in", "Pass123!", "student", "14:00-16:00", "Level 2"],
    ]
    for row in sample_rows:
        ws.append(row)

    # Style header row
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = cell.column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sample_student_import.xlsx"}
    )


# ---- Admin: Bulk Upload Users from Excel ----

@router.post("/users/bulk-upload")
def admin_bulk_upload_users(
    file: UploadFile = File(...),
    role: str = Form("student"),
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    """Upload an Excel (.xlsx/.xls) or CSV file to create users in bulk.

    Expected columns: roll_number, username, email, password, session_slot (or session), level_name (or level)
    """
    import io
    filename = file.filename or ""

    try:
        target_role = UserRole(role)
    except ValueError:
        target_role = UserRole.student

    try:
        raw = file.file.read()
    except Exception:
        raise HTTPException(400, "Could not read uploaded file")

    rows = []
    if filename.endswith(".csv"):
        import csv
        reader = csv.DictReader(io.StringIO(raw.decode("utf-8-sig")))
        for row in reader:
            rows.append({k.strip().lower(): v.strip() for k, v in row.items() if k})
    else:
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise HTTPException(500, "openpyxl not installed on server")
        wb = load_workbook(io.BytesIO(raw), read_only=True)
        ws = wb.active
        headers = [str(cell.value or "").strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = {headers[i]: str(row[i] or "").strip() for i in range(min(len(headers), len(row)))}
            if d.get("username") or d.get("roll_number") or d.get("email"):
                rows.append(d)

    # Pre-fetch levels mapping name -> id
    levels = db.query(Level).all()
    level_map = {l.name.strip().lower(): l.id for l in levels}

    created = 0
    skipped = 0
    for row in rows:
        roll = row.get("roll_number", "").strip() or row.get("roll_num", "").strip() or row.get("roll", "").strip()
        uname = row.get("username", "").strip() or roll or row.get("email", "").split("@")[0]
        email = row.get("email", "").strip()
        pwd = row.get("password", "").strip()
        slot = row.get("session_slot", "").strip() or row.get("session", "").strip() or row.get("slot", "").strip()
        level_input = row.get("level_name", "").strip() or row.get("level", "").strip()

        if not email:
            skipped += 1
            continue

        # Skip duplicates by roll_number or email for the SAME slot (NOT across different slots)
        slot_val = slot if slot else None
        filter_cond = (User.email == email) & (User.session_slot == slot_val)
        if roll:
            filter_cond = ((User.roll_number == roll) | (User.email == email)) & (User.session_slot == slot_val)

        if db.query(User).filter(filter_cond).first():
            skipped += 1
            continue

        # Match level
        matched_level_id = None
        if level_input:
            if level_input in level_map:
                matched_level_id = level_map[level_input]
            elif level_input.lower() in level_map:
                matched_level_id = level_map[level_input.lower()]
            else:
                lvl = db.query(Level).filter((Level.id == level_input) | (Level.name.ilike(f"%{level_input}%"))).first()
                if lvl:
                    matched_level_id = lvl.id

        hashed_pwd = auth_service.hash_password(pwd) if pwd else ""

        user = User(
            username=uname,
            email=email,
            hashed_password=hashed_pwd,
            role=target_role,
            is_active=True,
            roll_number=roll if roll else None,
            session_slot=slot if slot else None,
            level_id=matched_level_id,
        )
        db.add(user)
        created += 1

    db.commit()
    return {"message": f"{created} users created, {skipped} skipped", "created": created, "skipped": skipped}

